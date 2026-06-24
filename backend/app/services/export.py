"""Excel export service using openpyxl."""
from datetime import date
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import Lead, LeadSource, Service, LeadStage, User, Deal, FinanceTransaction, PayrollRecord

HEADER_FILL = PatternFill("solid", fgColor="4648D4")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _header(ws, columns: list[str]):
    for col, title in enumerate(columns, 1):
        cell = ws.cell(1, col, title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def export_leads_xlsx(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    source_id: Optional[int] = None,
    stage_id: Optional[int] = None,
) -> bytes:
    q = db.query(Lead)
    if date_from:
        q = q.filter(func.date(Lead.created_at) >= date_from)
    if date_to:
        q = q.filter(func.date(Lead.created_at) <= date_to)
    if source_id:
        q = q.filter(Lead.source_id == source_id)
    if stage_id:
        q = q.filter(Lead.stage_id == stage_id)
    leads = q.order_by(Lead.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лиды"

    cols = ["Дата", "Клиент", "Телефон", "Источник", "Услуга", "Сеттер", "Клоузер",
            "Этап", "Потенциал", "Факт", "Статус"]
    _header(ws, cols)

    for lead in leads:
        ws.append([
            lead.created_at.strftime("%d.%m.%Y") if lead.created_at else "",
            lead.client_name,
            lead.phone,
            lead.source.name if lead.source else "",
            lead.service.name if lead.service else "",
            lead.setter.name if lead.setter else "",
            lead.closer.name if lead.closer else "",
            lead.stage.name if lead.stage else "",
            lead.potential_amount,
            lead.actual_amount,
            "Активный" if lead.status == "active" else "Архив",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_finance_xlsx(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> bytes:
    q = db.query(FinanceTransaction)
    if date_from:
        q = q.filter(FinanceTransaction.date >= date_from)
    if date_to:
        q = q.filter(FinanceTransaction.date <= date_to)
    txs = q.order_by(FinanceTransaction.date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Финансы"

    cols = ["Дата", "Тип", "Категория", "Сумма", "Способ оплаты", "Ответственный", "Комментарий"]
    _header(ws, cols)

    for t in txs:
        ws.append([
            t.date.strftime("%d.%m.%Y") if t.date else "",
            "Доход" if t.type == "income" else "Расход",
            t.category or "",
            t.amount,
            t.payment_method or "",
            t.responsible.name if t.responsible else "",
            t.comment or "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_payroll_xlsx(
    db: Session,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> bytes:
    q = db.query(PayrollRecord)
    if date_from:
        q = q.filter(PayrollRecord.period_start >= date_from)
    if date_to:
        q = q.filter(PayrollRecord.period_end <= date_to)
    records = q.order_by(PayrollRecord.period_start.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Зарплаты"

    cols = ["Сотрудник", "Период с", "Период по", "Оклад", "Комиссия", "Бонус",
            "Штраф", "Итого", "Статус"]
    _header(ws, cols)

    for r in records:
        ws.append([
            r.employee.name if r.employee else "",
            r.period_start.strftime("%d.%m.%Y") if r.period_start else "",
            r.period_end.strftime("%d.%m.%Y") if r.period_end else "",
            r.base_salary,
            r.commission_amount,
            r.bonus_amount,
            r.penalty_amount,
            r.total_amount,
            "Выплачено" if r.status == "paid" else "Черновик",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
